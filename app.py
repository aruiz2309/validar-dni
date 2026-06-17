import streamlit as st
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re, unicodedata, time, random, io

st.set_page_config(page_title="Validador DNI Perú", page_icon="🪪", layout="centered")

st.title("🪪 Validador de DNI - Perú")
st.markdown("Sube tu Excel con **NOMBRE COMPLETO** y **NRO DNI**, valida y descarga el resultado.")

# ── Sidebar: configuración del token ─────────────────────────
with st.sidebar:
    st.header("⚙️ Configuración")
    st.markdown(
        "Esta app usa la API de **[apiperu.dev](https://apiperu.dev)** "
        "para consultar nombres por DNI desde la nube.\n\n"
       
    )
    token = st.text_input("🔑 Token API (apiperu.dev)", type="password", placeholder="Pega tu token aquí")
    if token:
        st.success("Token ingresado ✓")
    else:
        st.warning("Ingresa tu token para continuar")

# ── Utilidades ────────────────────────────────────────────────
def normalizar(t):
    if not t: return ""
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t.upper()).strip()

def similitud(a, b):
    pa = set(normalizar(a).split())
    pb = set(normalizar(b).split())
    return len(pa & pb) / len(pa) if pa else 0.0

# ── Consulta API apiperu.dev ──────────────────────────────────
def consultar_dni_api(dni, token):
    """
    Consulta el DNI usando la API de apiperu.dev
    Devuelve el nombre completo o cadena vacía si no encuentra.
    """
    try:
        r = requests.get(
            f"https://apiperu.dev/api/dni/{dni}",
            headers={
                "Authorization": f"Bearer {token}",
                "Accept": "application/json",
                "Content-Type": "application/json",
            },
            timeout=15,
        )
        if r.status_code == 200:
            j = r.json()
            # Estructura: {"success": true, "data": {"nombre_completo": "...", ...}}
            data = j.get("data", j)
            nc = data.get("nombre_completo", "")
            if nc:
                return nc.upper().strip()
            # Armar desde partes
            ap = data.get("apellido_paterno", "")
            am = data.get("apellido_materno", "")
            nb = data.get("nombres", "")
            if ap or nb:
                return re.sub(r"\s+", " ", f"{ap} {am} {nb}").strip().upper()
        elif r.status_code == 401:
            return "__TOKEN_INVALIDO__"
        elif r.status_code == 429:
            return "__LIMITE_ALCANZADO__"
    except Exception:
        pass
    return ""

def validar(nombre_arch, dni, token):
    nombre_web = consultar_dni_api(dni, token)

    if nombre_web == "__TOKEN_INVALIDO__":
        return {"nombre_web": "TOKEN INVÁLIDO", "estado": 0, "error": True}
    if nombre_web == "__LIMITE_ALCANZADO__":
        return {"nombre_web": "LÍMITE DE CONSULTAS ALCANZADO", "estado": 0, "error": True}
    if not nombre_web:
        return {"nombre_web": "NO ENCONTRADO", "estado": 0, "error": False}

    sim = similitud(nombre_arch, nombre_web)
    return {"nombre_web": nombre_web, "estado": 1 if sim >= 0.80 else 0, "error": False}

# ── Leer Excel ────────────────────────────────────────────────
def leer_excel(b):
    wb = openpyxl.load_workbook(io.BytesIO(b), read_only=True)
    ws = wb.active
    registros, vistos = [], set()
    fila_inicio = 2
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
        vals = [str(v).lower() for v in row if v]
        if any("nombre" in v or "dni" in v for v in vals):
            fila_inicio = i + 1
            break
    for row in ws.iter_rows(min_row=fila_inicio, values_only=True):
        nombre  = str(row[0]).strip() if row[0] else ""
        dni_raw = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        dni     = re.sub(r"\D", "", dni_raw).zfill(8)
        if not dni or not dni.replace("0", ""): continue
        if dni in vistos: continue
        vistos.add(dni)
        registros.append({"nombre": nombre, "dni": dni})
    return registros

# ── Generar Excel resultado ───────────────────────────────────
def generar_excel(registros, resultados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado"

    thin  = Side(style="thin", color="BFBFBF")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_cab   = PatternFill("solid", fgColor="1F4E79")
    fill_verde = PatternFill("solid", fgColor="E2EFDA")
    fill_rojo  = PatternFill("solid", fgColor="FCE4D6")
    fill_gris  = PatternFill("solid", fgColor="F2F2F2")
    fill_blanc = PatternFill("solid", fgColor="FFFFFF")
    font_cab   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    font_datos = Font(name="Arial", size=10)
    align_c    = Alignment(horizontal="center", vertical="center")
    align_l    = Alignment(horizontal="left",   vertical="center")

    ws.row_dimensions[1].height = 28
    for col, t in enumerate(["NOMBRE COMPLETO", "NRO DNI", "NOMBRE DE LA WEB", "ESTADO"], 1):
        c = ws.cell(row=1, column=col, value=t)
        c.font=font_cab; c.fill=fill_cab; c.border=borde; c.alignment=align_c

    for i, reg in enumerate(registros, 2):
        ws.row_dimensions[i].height = 18
        r      = resultados.get(reg["dni"], {"nombre_web": "NO PROCESADO", "estado": 0})
        estado = r["estado"]
        fill_f = fill_gris if i % 2 == 0 else fill_blanc
        fill_e = fill_verde if estado == 1 else fill_rojo

        for col, (val, fl, aln) in enumerate(zip(
            [reg["nombre"], reg["dni"], r["nombre_web"], estado],
            [fill_f, fill_f, fill_f, fill_e],
            [align_l, align_c, align_l, align_c]
        ), 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font=font_datos; c.fill=fl; c.border=borde; c.alignment=aln

    for col, ancho in enumerate([42, 14, 42, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── INTERFAZ ──────────────────────────────────────────────────
archivo = st.file_uploader(
    "📂 Sube tu Excel — columnas: **NOMBRE COMPLETO** | **NRO DNI**",
    type=["xlsx"]
)

if archivo and token:
    datos_bytes = archivo.read()
    registros   = leer_excel(datos_bytes)

    if not registros:
        st.error("❌ No se encontraron datos. Verifica que el Excel tenga las columnas correctas.")
    else:
        st.success(f"✅ **{len(registros)} DNIs únicos** listos para validar")

        col1, col2 = st.columns(2)
        col1.metric("Total DNIs", len(registros))
        mins = (len(registros) * 3) // 60
        segs = (len(registros) * 3) % 60
        col2.metric("Tiempo estimado", f"~{mins}m {segs}s")

        with st.expander("👁️ Vista previa"):
            for r in registros[:8]:
                st.markdown(f"• `{r['dni']}` → {r['nombre']}")
            if len(registros) > 8:
                st.caption(f"... y {len(registros)-8} más")

        st.warning("⚠️ No cierres esta pestaña mientras se procesa.")

        if st.button("🚀 Iniciar Validación", type="primary", use_container_width=True):
            resultados = {}
            barra      = st.progress(0, text="Iniciando...")
            estado_txt = st.empty()
            col_a, col_b, col_c = st.columns(3)
            cnt_ok  = col_a.empty()
            cnt_inc = col_b.empty()
            cnt_no  = col_c.empty()
            ok = inc = no = 0
            error_fatal = False

            for i, reg in enumerate(registros):
                dni    = reg["dni"]
                nombre = reg["nombre"]
                estado_txt.markdown(f"🔍 **{i+1}/{len(registros)}** — DNI: `{dni}` | `{nombre[:40]}`")

                r = validar(nombre, dni, token)
                resultados[dni] = r

                if r.get("error"):
                    estado_txt.error(f"❌ Error crítico: {r['nombre_web']}. Se detiene el proceso.")
                    error_fatal = True
                    break

                if r["estado"] == 1:           ok  += 1
                elif r["nombre_web"] == "NO ENCONTRADO": no += 1
                else:                          inc += 1

                cnt_ok.metric("✅ Estado 1",        ok)
                cnt_inc.metric("❌ Estado 0 (≠)",   inc)
                cnt_no.metric("⚠️ No encontrado",   no)
                barra.progress((i+1)/len(registros), text=f"Procesando {i+1} de {len(registros)}...")

                if i < len(registros) - 1:
                    time.sleep(random.uniform(0.5, 1.2))

            if not error_fatal:
                barra.progress(1.0, text="¡Completado!")
                estado_txt.success("✅ Validación finalizada")

            excel_buf = generar_excel(registros, resultados)
            st.download_button(
                label="📥 Descargar Excel con resultados",
                data=excel_buf,
                file_name="validacion_dni_resultado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

            st.markdown("---")
            st.subheader("📊 Resumen")
            total = len(resultados)
            c1, c2, c3 = st.columns(3)
            c1.metric("✅ Correctos (1)",   ok,  f"{round(ok/total*100) if total else 0}%")
            c2.metric("❌ Incorrectos (0)", inc, f"{round(inc/total*100) if total else 0}%")
            c3.metric("⚠️ No encontrados",  no,  f"{round(no/total*100) if total else 0}%")

            st.markdown("---")
            st.subheader("👁️ Vista previa del resultado")
            for reg in registros[:8]:
                r   = resultados.get(reg["dni"], {})
                nw  = r.get("nombre_web", "—")
                est = r.get("estado", 0)
                ico = "✅" if est == 1 else "❌"
                st.markdown(f"{ico} `{reg['dni']}` | **{reg['nombre']}** → `{nw}`")
            if len(registros) > 8:
                st.caption(f"... y {len(registros)-8} filas más en el Excel.")

elif archivo and not token:
    st.warning("⬅️ Ingresa tu token de apiperu.dev en el panel izquierdo para continuar.")
elif not archivo and token:
    st.info("📂 Sube tu archivo Excel para continuar.")
